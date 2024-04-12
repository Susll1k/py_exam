import sqlite3
import re
import time
from openpyxl import Workbook
import tabulate

connection = sqlite3.connect('database.db')

cursor = connection.cursor()

def create_table():
    cursor.execute('CREATE TABLE IF NOT EXISTS cinemas (id INTEGER PRIMARY KEY, name TEXT, addres TEXT)')
    cursor.execute('CREATE TABLE IF NOT EXISTS movies (id INTEGER PRIMARY KEY, name TEXT, genre TEXT, year INTEGER, description TEXT, rating REAL)')
    cursor.execute('CREATE TABLE IF NOT EXISTS afisha (id INTEGER PRIMARY KEY, movie_id INTEGER, cinema_id INTEGER, price INTEGER, date DATE, time TIME, capacity INTEGER)')
    cursor.execute('CREATE TABLE IF NOT EXISTS place (id INTEGER PRIMARY KEY, afisha_id INTEGER, room INTEGER, row INTEGER, seat INTEGER)')
    cursor.execute('CREATE TABLE IF NOT EXISTS ticket (id INTEGER PRIMARY KEY, name TEXT, phone TEXT, place_id INTEGER)')

create_table()





class Tiket():
    def __init__(self):
        pass
        cursor.execute(f"SELECT genre FROM movies")
        genres=cursor.fetchall()
        genresList=[]
        for i in range(len(genres)):
            genresList.append(genres[i][0])
        genresList=set(genresList)
        genresList=list(genresList)

        di=[]

        

        i=1
        while True:
            lst=[['Genre', 'Cinema', 'Movie']]
            for afisha in afishas:
                cursor.execute(f"SELECT genre FROM movies WHERE id = {afisha[1]}")
                genre=cursor.fetchall()
                if genre[0][0] == genresList[i-1]:
                    cursor.execute(f"SELECT name FROM cinemas WHERE id = {afisha[2]}")
                    cinema=cursor.fetchall()[0][0]
                    cursor.execute(f"SELECT name FROM movies WHERE id = {afisha[1]}")
                    movie=cursor.fetchall()[0][0]
                    lst.append([genresList[i-1], cinema, movie])
            if lst != [['Genre', 'Cinema', 'Movie']]:
                di.append(genresList[i-1])
                print(f'\n{di.index(f"{genresList[i-1]}")+1})\n')
                print(tabulate.tabulate(lst, headers='firstrow',tablefmt='grid'))

            lst=[['Genre', 'Cinema', 'Movie']]

            i+=1
            if i == len(genresList):
                break
        print(di)
        while True:
            try:
                genre=int(input(f'Выбери жанр (введи номер который указан рядом с названием нужного вам жанра от 1 до {len(di)}): '))
                if genre < 1:
                    print(f"Введи от 1 до {len(di)}")
                elif genre > len(di):
                    print(f"Введи от 1 до {len(di)}")
                else:
                    break
            except ValueError:
                print("Введи цифру")
        self.genre=di[genre-1]
    
    def show_cinemas(self, afishas):

        i=1
        while True:
            cursor.execute(f"SELECT name FROM cinemas WHERE id = {i}")
            print(f'\n{i}) {cursor.fetchall()[0][0]}\n')
            for afisha in afishas:
                if afisha[2] == i:
                    cursor.execute(f"SELECT name FROM movies WHERE id = {afisha[1]}")
                    print(f'{cursor.fetchall()[0][0]}, Дата и время: {afisha[4]}, {afisha[5]}, Стоимость: {afisha[3]}')
            i+=1
            if i == 6:
                break
        while True:
            try:
                cinema=int(input('Выбери кинотеатр (введи номер который указан рядом с названием нужного вам кинотеатра от 1 до 5): '))
                if cinema < 1:
                    print("Введи от 1 до 5")
                elif cinema > 5:
                    print("Введи от 1 до 5")
                else:
                    break
            except ValueError:
                print("Введи цифру")
        self.id_cinema=cinema
    
    def movie_selection(self, afishas):
        cursor.execute(f"SELECT name FROM cinemas WHERE id = {self.id_cinema}")
        print(f'\n{self.id_cinema}) {cursor.fetchall()[0][0]}')
        i=1
        list_id_movie=[]
        for afisha in afishas:
            if afisha[2] == self.id_cinema:
                cursor.execute(f"SELECT name FROM movies WHERE id = {afisha[1]}")
                name=cursor.fetchall()[0][0]
                print(f'{i}) {name}, Дата и время: {afisha[4]}, {afisha[5]}, Стоимость: {afisha[3]} \n')
                
                cursor.execute(f'SELECT genre FROM movies WHERE name = "{name}"')
                genre=cursor.fetchall()[0][0]

                cursor.execute(f'SELECT year FROM movies WHERE name = "{name}"')
                year=cursor.fetchall()[0][0]

                cursor.execute(f'SELECT description FROM movies WHERE name = "{name}"')
                description=cursor.fetchall()[0][0]

                cursor.execute(f'SELECT rating FROM movies WHERE name = "{name}"')
                rating=cursor.fetchall()[0][0]

                cursor.execute(f'SELECT id FROM movies WHERE name = "{name}"')
                id=cursor.fetchall()[0][0]
                list_id_movie.append(id)



                print(f'Description:\nGenre: {genre}\nYear: {year}\nDescription: {description}\nRating: {rating}\n')
                i+=1
        while True:
            try:
                movie=int(input(f'Выбери фильм (введи номер который указан рядом с названием нужного вам фильма от 1 до {i-1}): '))
                if movie < 1:
                    print(f"Введи от 1 до {i-1}")
                elif movie > i-1:
                    print(f"Введи от 1 до {i-1}")
                else:
                    break
            except ValueError:
                print("Введи цифру")
        self.id_movie=list_id_movie[movie-1]
        cursor.execute(f"SELECT id FROM afisha WHERE movie_id={self.id_movie} AND cinema_id = {self.id_cinema}")
        self.afisha_id=cursor.fetchall()[0][0]

    def place(self):
        cursor.execute(f"SELECT * FROM place WHERE afisha_id = {self.afisha_id}")
        places=cursor.fetchall()
        print(places)
        print(f'Зал: {places[0][2]}')
        i=1
        for place in places:
            print(f"{i}) Ряд: {place[3]}, Место: {place[4]}")
            i+=1
        while True:
            try:
                place=int(input(f'Выбери место (введи номер который указан рядом с понравившимся вам места от 1 до {len(places)}): '))
                if place < 1:
                    print(f"Введи от 1 до {len(places)}")
                elif place > len(places):
                    print(f"Введи от 1 до {len(places)}")
                else:
                    break
            except ValueError:
                print("Введи цифру")
        self.id_place=places[place-1][0]
        
        

class Cheakout(Tiket):
    def __init__(self, name):
        self.name=name
    def add_db(self):
        while True:
            number=input('Номер телефона по которому будет произведена оплата. Напиши в таком формате(+7-ххх-ххх-хх-хх)')
            print("Проверка....")
            time.sleep(5)
            if re.match(r"[+][7]-[\d]{3}-[\d]{3}-[\d]{2}-[\d]{2}", number):
                print('Проверка прошла успешно!')
                

                break
            else:
                print('Ошибка')
        cursor.execute(f'INSERT INTO ticket (name, phone, place_id) VALUES ("{self.name}", "{number}", {self.id_place})')
        connection.commit()

    def creating_receipt(self):
        cursor.execute(f"SELECT * FROM ticket")
        tickets=cursor.fetchall()
    
        


        wb = Workbook()
        ws = wb.active

        ws['A1'] = "Name"
        ws['B1'] = "Phone"
        ws['C1'] = "Place"
        ws['D1'] = "Data"
        ws['E1'] = "Time"
        ws['F1'] = "Cinema"
        ws['G1'] = "Movie"

        price=0

        for i in range(len(tickets)):
            ws[f'A{i+2}'] = tickets[i][1]
            ws[f'B{i+2}'] = tickets[i][2]
            cursor.execute(f"SELECT * FROM place WHERE id = {tickets[i][3]}")
            places=cursor.fetchall()
            ws[f'C{i+2}'] = f'Зал: {places[0][2]}\nРяд: {places[0][3]}, Место: {places[0][4]}'
            cursor.execute(f"SELECT * FROM afisha WHERE id = {places[0][1]}")
            afishas=cursor.fetchall()
            ws[f'D{i+2}'] = f'{afishas[0][4]}'
            ws[f'E{i+2}'] = f'{afishas[0][5]}'
            cursor.execute(f"SELECT * FROM cinemas WHERE id = {afishas[0][2]}")
            cinema=cursor.fetchall()
            ws[f'F{i+2}'] = f'{cinema[0][1]}'
            cursor.execute(f"SELECT * FROM movies WHERE id = {afishas[0][1]}")
            movie=cursor.fetchall()
            ws[f'G{i+2}'] = f'{movie[0][1]}'

            price+=afishas[0][3]
            id=i
        
        ws[f'A{id+3}'] = f'Total price: {price}'






        wb.save("cheak.xlsx")



print("\nПриветствую вас на TikeSusl!")
print('Здесь есть множество различных фильмов')

cursor.execute("SELECT * FROM afisha")
afishas = cursor.fetchall()


while True:

    while True:
        try:
            name=input('\n  Напиши свое имя: ')
            if name.isdigit():
                raise ValueError()
            break
        except ValueError:
            print('Ошибка!')
    cheak=Cheakout(name)
    
    
    cheak.show_cinemas(afishas)
    cheak.movie_selection(afishas)

    print('Теперь выбери место!')
    print('Вот все свободные места.')
    cheak.place()

        
    cheak.add_db()
    cheak.creating_receipt()
    

    while True:
        try:
            task=int(input('\nХотите выбрать еще билет(если да то напиши(1), если нет то напиши(2)): '))
            if task >2:
                raise ValueError
            elif task <1:
                raise ValueError
            break
        
        except ValueError:
            print('Ошибка')
    if task == 2:
        break 



    



    




    

